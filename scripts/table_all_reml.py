#!/usr/bin/env python3
"""
Parse REML files and generate Excel report with heritability information.
"""

import argparse
import glob
import os
import re
from pathlib import Path


def parse_filename(filename):
    """
    Extract phenotype and Type from filename.
    Format: ${phenotype}.${Type}.reml
    Example: GYP_BLUP.SV.reml -> phenotype='GYP_BLUP', type='SV'
    """
    basename = os.path.basename(filename)
    # Remove .reml extension
    name_without_ext = basename.replace('.reml', '')
    
    # Split by last dot to separate phenotype and Type
    # Handle cases like SNP_INDEL_SV where Type contains underscores
    parts = name_without_ext.rsplit('.', 1)
    if len(parts) == 2:
        phenotype, type_val = parts
        return phenotype, type_val
    else:
        # Fallback: try to find pattern
        match = re.match(r'^(.+)\.([^.]+)$', name_without_ext)
        if match:
            return match.group(1), match.group(2)
        return name_without_ext, 'UNKNOWN'


def parse_reml_file(filepath):
    """
    Parse a .reml file and extract heritability information.
    
    Returns:
        dict with keys: converged, her_k1, her_k2, her_k3, her_all
        Each her_* value is a list: [Heritability, SE, Size, Mega_Intensity, SE]
    """
    result = {
        'converged': None,
        'her_k1': None,
        'her_k2': None,
        'her_k3': None,
        'her_all': None
    }
    
    with open(filepath, 'r') as f:
        lines = f.readlines()
    
    # Find Converged status
    for line in lines:
        if line.startswith('Converged'):
            parts = line.strip().split()
            if len(parts) >= 2:
                result['converged'] = parts[1]
            break
    
    # Find Component header line
    component_start_idx = None
    for i, line in enumerate(lines):
        if line.startswith('Component'):
            component_start_idx = i
            break
    
    if component_start_idx is None:
        return result
    
    # Parse Her_* lines
    for i in range(component_start_idx + 1, len(lines)):
        line = lines[i].strip()
        if not line:
            continue
        
        parts = line.split()
        if len(parts) < 2:
            continue
        
        component = parts[0]
        
        # Extract values (skip component name)
        values = parts[1:]
        
        # Parse values, handling NA
        parsed_values = []
        for val in values[:5]:  # Take first 5 values: Heritability, SE, Size, Mega_Intensity, SE
            if val == 'NA':
                parsed_values.append(None)
            else:
                try:
                    parsed_values.append(float(val))
                except ValueError:
                    parsed_values.append(None)
        
        # Ensure we have 5 values (pad with None if needed)
        while len(parsed_values) < 5:
            parsed_values.append(None)
        
        # Store based on component name
        if component == 'Her_K1':
            result['her_k1'] = parsed_values
        elif component == 'Her_K2':
            result['her_k2'] = parsed_values
        elif component == 'Her_K3':
            result['her_k3'] = parsed_values
        elif component == 'Her_All':
            result['her_all'] = parsed_values
    
    return result


def main():
    parser = argparse.ArgumentParser(
        description='Parse REML files and generate Excel report with heritability information.'
    )
    parser.add_argument(
        '--pattern',
        required=True,
        help='Glob pattern for .reml files (e.g., "dir/*.reml")'
    )
    parser.add_argument(
        '-o', '--output',
        required=True,
        help='Output Excel file path'
    )
    
    args = parser.parse_args()
    
    # Find all matching .reml files
    reml_files = glob.glob(args.pattern)
    
    if not reml_files:
        print(f"Warning: No files found matching pattern '{args.pattern}'")
        return
    
    print(f"Found {len(reml_files)} .reml files")
    
    # Parse all files
    summary_data_dict = {}  # {phenotype: {type: {converged, her_k1, her_k2, her_k3, her_all}}}
    detailed_data = []
    
    for filepath in sorted(reml_files):
        phenotype, type_val = parse_filename(filepath)
        parsed = parse_reml_file(filepath)
        
        # Store summary data by phenotype and type
        if phenotype not in summary_data_dict:
            summary_data_dict[phenotype] = {}
        
        summary_data_dict[phenotype][type_val] = {
            'converged': parsed['converged'],
            'her_k1': parsed['her_k1'][0] if parsed['her_k1'] and parsed['her_k1'][0] is not None else None,
            'her_k2': parsed['her_k2'][0] if parsed['her_k2'] and parsed['her_k2'][0] is not None else None,
            'her_k3': parsed['her_k3'][0] if parsed['her_k3'] and parsed['her_k3'][0] is not None else None,
            'her_all': parsed['her_all'][0] if parsed['her_all'] and parsed['her_all'][0] is not None else None,
        }
        
        # Prepare detailed rows (all components with all fields)
        components = [
            ('Her_K1', parsed['her_k1']),
            ('Her_K2', parsed['her_k2']),
            ('Her_K3', parsed['her_k3']),
            ('Her_All', parsed['her_all']),
        ]
        
        for component_name, component_data in components:
            if component_data:
                detailed_row = {
                    'Phenotype': phenotype,
                    'Type': type_val,
                    'Component': component_name,
                    'Converged': parsed['converged'],
                    'Heritability': component_data[0],
                    'SE': component_data[1],
                    'Size': component_data[2],
                    'Mega_Intensity': component_data[3],
                    'SE_2': component_data[4],  # Second SE column
                }
                detailed_data.append(detailed_row)
    
    # Write to Excel using openpyxl for merged cells
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        import pandas as pd
    except ImportError:
        print("Error: openpyxl and pandas are required. Please install them with: pip install pandas openpyxl")
        return
    
    # Collect all unique types
    all_types = set()
    for phenotype_data in summary_data_dict.values():
        all_types.update(phenotype_data.keys())
    all_types = sorted(all_types)
    
    # Create workbook
    wb = Workbook()
    
    # Create summary sheet with merged cells
    ws_summary = wb.active
    ws_summary.title = 'Heritability Summary'
    
    # Column headers: each Type has 5 columns (Her_K1, Her_K2, Her_K3, Her_All, Converged)
    col_headers = ['Her_K1', 'Her_K2', 'Her_K3', 'Her_All', 'Converged']
    
    # Write header row 1 (merged cells for Type names)
    ws_summary.cell(row=1, column=1, value='Phenotype')
    col_idx = 2
    type_start_cols = {}  # Track where each type starts
    
    for type_val in all_types:
        type_start_cols[type_val] = col_idx
        # Merge cells for Type name (5 columns)
        ws_summary.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 4)
        cell = ws_summary.cell(row=1, column=col_idx, value=type_val)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        col_idx += 5
    
    # Write header row 2 (sub-columns for each Type)
    ws_summary.cell(row=2, column=1, value='Phenotype').font = Font(bold=True)
    col_idx = 2
    for type_val in all_types:
        for header in col_headers:
            cell = ws_summary.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            col_idx += 1
    
    # Write data rows
    row_idx = 3
    for phenotype in sorted(summary_data_dict.keys()):
        ws_summary.cell(row=row_idx, column=1, value=phenotype)
        col_idx = 2
        for type_val in all_types:
            if type_val in summary_data_dict[phenotype]:
                data = summary_data_dict[phenotype][type_val]
                ws_summary.cell(row=row_idx, column=col_idx, value=data['her_k1'])
                ws_summary.cell(row=row_idx, column=col_idx + 1, value=data['her_k2'])
                ws_summary.cell(row=row_idx, column=col_idx + 2, value=data['her_k3'])
                ws_summary.cell(row=row_idx, column=col_idx + 3, value=data['her_all'])
                ws_summary.cell(row=row_idx, column=col_idx + 4, value=data['converged'])
            else:
                # Fill with None/empty for missing data
                for i in range(5):
                    ws_summary.cell(row=row_idx, column=col_idx + i, value=None)
            col_idx += 5
        row_idx += 1
    
    # Create detailed sheet using pandas
    ws_detailed = wb.create_sheet('Detailed Information')
    df_detailed = pd.DataFrame(detailed_data)
    
    # Write headers for detailed sheet
    for c_idx, col_name in enumerate(df_detailed.columns, start=1):
        cell = ws_detailed.cell(row=1, column=c_idx, value=col_name)
        cell.font = Font(bold=True)
    
    # Write detailed data to sheet (starting from row 2)
    for r_idx, row in enumerate(df_detailed.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws_detailed.cell(row=r_idx, column=c_idx, value=value)
    
    # Save workbook
    wb.save(args.output)
    
    print(f"Successfully wrote results to {args.output}")
    print(f"  - Summary sheet: {len(summary_data_dict)} rows (with merged headers)")
    print(f"  - Detailed sheet: {len(df_detailed)} rows")


if __name__ == '__main__':
    main()

